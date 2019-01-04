# -*- coding: utf-8 -*-
# @Time    : 2018/12/5 20:08
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py

#
# 要求：
# 1:根据老师上课讲解的测试用例参数化，去写单元测试用例以及用到超继承
# 2：把所有的测试用例数据存到Excel中，然后写一个类do_excel，完成测试数据的读取
# 3：利用do_excel类读取到的测试数据 传给1203课堂上讲解的test_suite，完成测试用例的加载。
# 4：要求：登录3条用例 充值3条用例
# 5：请给用例加上断言 以及异常处理，断言可以用code 也可以用msg或者是用整段信息都OK，最后要有测试报告出具
from openpyxl import load_workbook

class DoExcel:

    def get_data(self,file_name,sheet_name,button):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]

        test_data=[]#存储所有行的数据
        for i in range(2,sheet.max_row+1):
            row_data={}
            #把每一行每个单元格的数据以key:value的形式存到row_data这个字典里面
            row_data['case_id']=sheet.cell(i,1).value
            row_data['title']=sheet.cell(i,2).value
            row_data['url']=sheet.cell(i,3).value
            row_data['http_method']=sheet.cell(i,4).value
            row_data['param']=sheet.cell(i,5).value
            row_data['expected']=sheet.cell(i,6).value

            #这里是把每一行的数据添加到test_data这里列表里面
            test_data.append(row_data)

        final_data=[]#最后的测试数据
        if button=='all':
            final_data=test_data
        else:#这里去拓展？？ 怎么去读取button里面的值 列表里面的case_id
            for item in test_data:#遍历test_data里面的数据
                if item['case_id'] in eval(button):#判断 case_id 在不在我button里面
                    #如果在的话 就把该条数据加到final_data里面去  如果不在的话 就不加
                    final_data.append(item)
        return final_data

#测试代码
if __name__ == '__main__':
    from homework.api_auto_3.common.read_config import ReadConfig

    button=ReadConfig().read_config('case.conf','CASE','button')
    test_data=DoExcel().get_data('test_cases.xlsx','test_data',button)
    print(test_data)