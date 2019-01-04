#!/usr/bin/env python 
# -*- coding:utf-8 -*-
# @Time :2018/12/4 20:23
''''#xlrd xlwt xlwriter  都是可以操作读写'''
from openpyxl import load_workbook  #可以对execl进行读写
'''打开文件'''
# workbook=load_workbook('python13.xlsx')  #相对路径，在跟自己同级的目录下
'''定位表单'''
# sheet=workbook['Sheet1']#字典类似    #返回表单对象
# sheet_1=workbook.get_sheet_by_name('Sheet1')  #已经过时，会警告信息
'''终于可以读大佬的名字'''
# dalao=sheet.cell(4,4).value
'''从1 开始的，读数据行列值，是从1 开始读的'''
# print(dalao)  #打印王波
# dada=sheet.cell(3,3).value
# print(dada)   #打印叶子
'''如何获取最大的行列值'''
# max_row=sheet.max_row
# max_col=sheet.max_column
# print('行',max_row)
# print('列',max_col)

'''怎么写'''
# sheet.cell(4,4).value='Mrs lu'#等同于==赋值
# print(sheet.cell(4,4).value)
# workbook.save('python13.xlsx')
# workbook.save('python14.xlsx')  #等同于另存为


'''怎么去新建execl'''
# from openpyxl import Workbook
# wb=Workbook()#创建对象
# wb.save('python14.xlsx')


'''如何把每一换行的数据读取存到一个字典里面，
所有行的数据以字典格式存在一个列表中'''
wb=load_workbook('python13.xlsx')
sheet=wb['Sheet1']
test_data=[]
for i in range (1,sheet.max_row+1):
    sub_data={}
    sub_data['A']=sheet.cell(i,1).value
    sub_data['B']=sheet.cell(i,2).value
    sub_data['C']=sheet.cell(i,3).value
    sub_data['D']=sheet.cell(i,4).value
    sub_data['E']=sheet.cell(i,5).value
    test_data.append(sub_data)
print(test_data)

'''存在execl里面的数据类型，整数、浮点数  读取出来的还是整数、浮点数
字符串字典 列表等其他数据类型  读出来的都是字符串'''

'''判断整形类型   读出来是整形'''
workbook=load_workbook('python13.xlsx')
sheet=workbook['Sheet1']
dadada=sheet.cell(1,2).value
print(dadada)
print(type(dadada))

'''判断浮点数类型  读出来的是浮点型'''
workbook=load_workbook('python13.xlsx')
sheet=workbook['Sheet1']
dadada=sheet.cell(1,3).value
print(dadada)
print(type(dadada))

'''判断字典类型  读出来的是字符串'''
workbook=load_workbook('python13.xlsx')
sheet=workbook['Sheet1']
dadada=sheet.cell(2,2).value
print(dadada)
print(type(dadada))


'''判断列表类型  读出来的是字符串'''
workbook=load_workbook('python13.xlsx')
sheet=workbook['Sheet1']
dadada=sheet.cell(2,3).value
print(dadada)
print(type(dadada))

