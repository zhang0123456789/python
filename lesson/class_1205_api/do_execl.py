#!/usr/bin/env python 
# -*- coding:utf-8 -*-
# @Time :2018/12/4 20:23
''''#xlrd xlwt xlwriter  都是可以操作读写'''
from openpyxl import load_workbook  #可以对execl进行读写

'''断言异常处理  读取表格数据 生成测试报告 test_api12.html'''
# class DoExecl:
#     def get_data(self,file_name,sheet_name):
#         wb=load_workbook('test_cases.xlsx')
#         sheet=wb['test_data']
#
#         test_data=[]
#         for i in range(2,sheet.max_row+1):#按行获取
#             row_data={}  #每一行的数据为空数据
#
#             '''把每一行每个单元格的数据以key:value的形式存到row_data这个知道里面'''
#             row_data['case_id']=sheet.cell(i,1).value#行由i确定
#             row_data['title']=sheet.cell(i,2).value
#             row_data['url']=sheet.cell(i, 3).value
#             row_data['http_method']=sheet.cell(i, 4).value
#             row_data['param']=sheet.cell(i, 5).value
#             row_data['expected']=sheet.cell(i, 6).value
#
#             test_data.append(row_data)#把每一行的数据添加到test_data这个列表里面
#
#
#         return test_data

'''测试代码'''
# if __name__ == '__main__':#测试代码写在这个里面，只有在这个main里面，才会被执行
#     test_data=DoExecl().get_data('test_cases.xlsx','test_data')
#     print(test_data)


'''eval的作用
a=1001
print(type(eval(a)))'''


'''断言异常处理 根据配置文件  读取表格数据 生成测试报告 test_api13.html'''
from openpyxl import load_workbook  #可以对execl进行读写
class DoExecl:
    def get_data(self,file_name,sheet_name,button):
        wb=load_workbook('test_cases.xlsx')#打开文件
        sheet=wb['test_data']#定位表单

        test_data=[]
        for i in range(2,sheet.max_row+1):#按行获取
            row_data={}  #每一行的数据为空数据

            '''把每一行每个单元格的数据以key:value的形式存到row_data这个知道里面'''
            row_data['case_id']=sheet.cell(i,1).value#行由i确定
            row_data['title']=sheet.cell(i,2).value
            row_data['url']=sheet.cell(i, 3).value
            row_data['http_method']=sheet.cell(i, 4).value
            row_data['param']=sheet.cell(i, 5).value
            row_data['expected']=sheet.cell(i, 6).value

            test_data.append(row_data)#把每一行的数据添加到test_data这个列表里面
        final_data=[]#最后的测试数据
        if button=='all':
            final_data=test_data
        else:
            for item in test_data: #遍历test_data里面的数据
                if item['case_id'] in eval(button):#判断case_id在不在button里面
                    final_data.append(item)

        return final_data

'''测试代码'''
if __name__ == '__main__':#测试代码写在这个里面，只有在这个main里面，才会被执行
    test_data=DoExecl().get_data('test_cases.xlsx','test_data')
    print(test_data)
